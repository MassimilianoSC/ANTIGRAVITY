import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from io import BytesIO
import os
import sys

# Add app to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.main import app

client = TestClient(app)

def create_test_xlsx():
    """Crea un file Excel di test in memoria."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Aggiungi header
    ws['A1'] = "Nome"
    ws['B1'] = "Età"
    ws['C1'] = "Città"
    
    # Aggiungi dati
    ws['A2'] = "Mario"
    ws['B2'] = 30
    ws['C2'] = "Roma"
    
    ws['A3'] = "Laura"
    ws['B3'] = 25
    ws['C3'] = "Milano"
    
    ws['A4'] = "Giuseppe"
    ws['B4'] = 35
    ws['C4'] = "Napoli"
    
    # Salva in BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    
    return xlsx_file

def test_health_endpoint():
    """Test health check endpoint."""
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}

def test_upload_xlsx():
    """Test upload di un file Excel."""
    xlsx_file = create_test_xlsx()
    
    response = client.post(
        "/api/datasets",
        files={"file": ("test.xlsx", xlsx_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    
    assert response.status_code == 200
    data = response.json()
    
    assert "id" in data
    assert data["filename"] == "test.xlsx"
    assert "sha256" in data
    assert data["sheet_count"] >= 1
    
    # Test GET dataset detail
    dataset_id = data["id"]
    response = client.get(f"/api/datasets/{dataset_id}")
    
    assert response.status_code == 200
    detail = response.json()
    
    assert "dataset" in detail
    assert "sheets" in detail
    assert len(detail["sheets"]) >= 1
    
    # Test preview grid
    sheet_name = detail["sheets"][0]["sheet_name"]
    response = client.get(
        f"/api/datasets/{dataset_id}/sheets/{sheet_name}/preview",
        params={"mode": "grid"}
    )
    
    assert response.status_code == 200
    preview = response.json()
    assert preview["mode"] == "grid"
    assert "data" in preview
    assert len(preview["data"]) > 0

def test_list_datasets():
    """Test lista datasets."""
    response = client.get("/api/datasets")
    assert response.status_code == 200
    assert isinstance(response.json(), list)

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
