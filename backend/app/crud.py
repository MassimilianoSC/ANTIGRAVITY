import hashlib
import uuid
import os
import json
from pathlib import Path
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from fastapi import UploadFile

from . import models, schemas
from .table_detection import analyze_sheet_structure

UPLOAD_DIR = Path("storage/uploads")

def get_datasets(db: Session, skip: int = 0, limit: int = 100) -> List[models.Dataset]:
    return db.query(models.Dataset).offset(skip).limit(limit).all()

def get_dataset(db: Session, dataset_id: str) -> models.Dataset:
    return db.query(models.Dataset).filter(models.Dataset.id == dataset_id).first()

def get_sheets_by_dataset(db: Session, dataset_id: str) -> List[models.Sheet]:
    return db.query(models.Sheet).filter(models.Sheet.dataset_id == dataset_id).all()

def get_sheet(db: Session, dataset_id: str, sheet_name: str) -> models.Sheet:
    return db.query(models.Sheet).filter(
        models.Sheet.dataset_id == dataset_id,
        models.Sheet.sheet_name == sheet_name
    ).first()

async def create_dataset(db: Session, file: UploadFile) -> models.Dataset:
    """Crea dataset da file xlsx caricato."""
    
    # Genera ID univoco
    dataset_id = str(uuid.uuid4())
    
    # Crea directory per il dataset
    dataset_dir = UPLOAD_DIR / dataset_id
    dataset_dir.mkdir(parents=True, exist_ok=True)
    
    # Salva file originale
    file_path = dataset_dir / "original.xlsx"
    content = await file.read()
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Calcola SHA256
    sha256_hash = hashlib.sha256(content).hexdigest()
    
    # Crea record Dataset
    dataset = models.Dataset(
        id=dataset_id,
        filename=file.filename,
        sha256=sha256_hash,
        file_path=str(file_path),
        file_size=len(content)
    )
    
    db.add(dataset)
    db.commit()
    db.refresh(dataset)
    
    # Processa il file Excel
    process_excel_file(db, dataset_id, file_path)
    
    return dataset

def process_excel_file(db: Session, dataset_id: str, file_path: Path):
    """Processa file Excel e salva sheet + celle."""
    
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Calcola dimensioni
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        # Conta merged cells (best effort)
        merged_count = 0
        try:
            if hasattr(ws, 'merged_cells'):
                merged_count = len(ws.merged_cells.ranges)
        except:
            merged_count = 0
        
        # Estrai celle non vuote
        cells_data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value is not None and str(value).strip():
                    value_text = str(value).strip()[:2000]
                    cells_data.append({
                        "dataset_id": dataset_id,
                        "sheet_name": sheet_name,
                        "row": row_idx,
                        "col": col_idx,
                        "value_text": value_text
                    })
        
        # Batch insert celle (1000 alla volta)
        batch_size = 1000
        for i in range(0, len(cells_data), batch_size):
            batch = cells_data[i:i+batch_size]
            db.bulk_insert_mappings(models.Cell, batch)
            db.commit()
        
        # Analizza struttura tabella
        cells_objects = db.query(models.Cell).filter(
            models.Cell.dataset_id == dataset_id,
            models.Cell.sheet_name == sheet_name
        ).all()
        
        analysis_json = analyze_sheet_structure(cells_objects)
        
        # Crea record Sheet
        sheet = models.Sheet(
            dataset_id=dataset_id,
            sheet_name=sheet_name,
            n_rows=max_row,
            n_cols=max_col,
            merged_cells_count=merged_count,
            analysis_json=analysis_json
        )
        
        db.add(sheet)
        db.commit()
    
    wb.close()

def get_grid_preview(
    db: Session,
    dataset_id: str,
    sheet_name: str,
    row_start: int = 1,
    row_end: int = 50,
    col_start: int = 1,
    col_end: int = 20
) -> Dict[str, Any]:
    """Restituisce preview in modalità griglia."""
    
    # Query celle nel range
    cells = db.query(models.Cell).filter(
        models.Cell.dataset_id == dataset_id,
        models.Cell.sheet_name == sheet_name,
        models.Cell.row >= row_start,
        models.Cell.row <= row_end,
        models.Cell.col >= col_start,
        models.Cell.col <= col_end
    ).all()
    
    # Crea mappa celle
    cell_map = {(c.row, c.col): c.value_text for c in cells}
    
    # Costruisci matrice 2D
    data = []
    for row in range(row_start, row_end + 1):
        row_data = []
        for col in range(col_start, col_end + 1):
            row_data.append(cell_map.get((row, col), ""))
        data.append(row_data)
    
    return {
        "mode": "grid",
        "data": data,
        "dimensions": {
            "row_start": row_start,
            "row_end": row_end,
            "col_start": col_start,
            "col_end": col_end
        }
    }

def get_table_preview(
    db: Session,
    dataset_id: str,
    sheet_name: str,
    candidate_idx: int = 0,
    raw_values: bool = False
) -> Dict[str, Any]:
    """Restituisce preview in modalità tabella."""
    
    # Carica analysis JSON
    sheet = get_sheet(db, dataset_id, sheet_name)
    if not sheet or not sheet.analysis_json:
        return {
            "mode": "table",
            "headers": [],
            "rows": [],
            "confidence": "low",
            "score": 0.0
        }
    
    analysis = json.loads(sheet.analysis_json)
    candidates = analysis.get("candidates", [])
    
    if candidate_idx >= len(candidates):
        return {
            "mode": "table",
            "headers": [],
            "rows": [],
            "confidence": "low",
            "score": 0.0
        }
    
    candidate = candidates[candidate_idx]
    row_start = candidate["row_start"]
    row_end = min(candidate["row_end"], candidate["row_start"] + 100)
    col_start = candidate["col_start"]
    col_end = candidate["col_end"]
    header_row = candidate.get("header_row", row_start)
    
    # Query celle nel rettangolo
    cells = db.query(models.Cell).filter(
        models.Cell.dataset_id == dataset_id,
        models.Cell.sheet_name == sheet_name,
        models.Cell.row >= row_start,
        models.Cell.row <= row_end,
        models.Cell.col >= col_start,
        models.Cell.col <= col_end
    ).all()
    
    cell_map = {(c.row, c.col): c.value_text for c in cells}
    
    # Estrai headers
    headers = []
    date_columns = []  # Indici colonne che contengono "DATA" nell'header
    for col_idx, col in enumerate(range(col_start, col_end + 1)):
        header_val = cell_map.get((header_row, col), f"Col{col}")
        headers.append(header_val)
        # Rileva colonne data
        if "DATA" in header_val.upper():
            date_columns.append(col_idx)
    
    # Estrai righe dati (fino a 50)
    rows = []
    data_start_row = header_row + 1
    for row in range(data_start_row, min(data_start_row + 50, row_end + 1)):
        row_data = []
        for col_idx, col in enumerate(range(col_start, col_end + 1)):
            value = cell_map.get((row, col), "")
            
            # Formatta date se richiesto e se la colonna è una colonna data
            if not raw_values and col_idx in date_columns and value:
                formatted_value = format_excel_date(value)
                row_data.append(formatted_value)
            else:
                row_data.append(value)
        rows.append(row_data)
    
    return {
        "mode": "table",
        "headers": headers,
        "rows": rows,
        "confidence": candidate.get("confidence", "low"),
        "score": candidate.get("score", 0.0)
    }

def format_excel_date(value: str) -> str:
    """Converte seriale Excel in data dd/mm/yyyy se possibile."""
    try:
        # Prova a convertire in numero
        serial = float(value)
        
        # Excel date serials: giorni dal 1900-01-01
        # Valori ragionevoli: tra 1 (1900) e 60000 (circa anno 2064)
        if 1 <= serial <= 60000:
            from datetime import datetime, timedelta
            
            # Separa parte intera (giorni) e decimale (frazione di giorno = ora)
            days = int(serial)
            time_fraction = serial - days
            
            # Excel considera erroneamente il 1900 come bisestile
            if days > 59:
                days -= 1
            
            base_date = datetime(1899, 12, 30)
            excel_date = base_date + timedelta(days=days)
            
            # Aggiungi la componente temporale se presente
            if time_fraction > 0:
                excel_date += timedelta(days=time_fraction)
            
            # Formatta in base alla presenza di orario
            # Se ora è 00:00:00 => solo data
            if excel_date.hour == 0 and excel_date.minute == 0 and excel_date.second == 0:
                return excel_date.strftime("%d/%m/%Y")
            else:
                # Include ora e minuti (senza secondi)
                return excel_date.strftime("%d/%m/%Y %H:%M")
        else:
            return value
    except (ValueError, TypeError, OverflowError):
        # Non è un numero o non può essere convertito
        return value
